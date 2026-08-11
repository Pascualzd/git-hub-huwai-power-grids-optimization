#!/usr/bin/env python3
"""Build the Oʻahu dispatchable fleet from EIA-860/923 and HECO IGP inputs."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EIA860 = ROOT / "data/raw/eia8602024/3_1_Generator_Y2024.xlsx"
EIA923 = ROOT / "data/raw/eia9232024/EIA923_Schedules_2_3_4_5_M_12_2024_Final.xlsx"
HECO_WB3 = ROOT / "data/raw/heco_20220331_final_oahu_inputs_workbook_3_revised.xlsx"
OUTPUT = ROOT / "data/processed/oahu_generators.csv"
EGRID_OUTPUT = ROOT / "data/processed/oahu_egrid_emissions.csv"

# Utility and contracted firm resources that serve the Oʻahu system. Emergency airport
# generators and refinery self-generation are intentionally excluded from the normal fleet.
PLANT_CODES = {765, 766, 10334, 54646, 56329, 60328}

FUEL_PRICE_COLUMN = {
    "LSFO": 3,
    "Diesel": 4,
    "ULSD_CIP": 5,
    "ULSD_SGS": 6,
}

CO2_TONNES_PER_MMBTU = {
    "LSFO": 0.083063247278,
    "Diesel": 0.081806613878,
    "ULSD_CIP": 0.081806613878,
    "ULSD_SGS": 0.081806613878,
}

VOM_USD_PER_MWH = {
    "steam": 5.0,
    "combustion_turbine": 4.5,
    "combined_cycle": 4.0,
    "internal_combustion": 5.0,
    "municipal_waste": 10.0,
}


def numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0.0)


def eia923_heat_rates() -> dict[tuple[int, str], float]:
    generation = pd.read_excel(EIA923, sheet_name="Page 1 Generation and Fuel Data", header=5)
    generation.columns = [str(column).replace("\n", " ") for column in generation.columns]
    generation = generation[generation["Plant Id"].isin(PLANT_CODES)].copy()
    generation["fuel_mmbtu"] = numeric(generation["Elec Fuel Consumption MMBtu"])
    generation["net_mwh"] = numeric(generation["Net Generation (Megawatthours)"])

    rates: dict[tuple[int, str], float] = {}
    for plant_id, plant in generation.groupby("Plant Id"):
        plant_id = int(plant_id)
        if plant_id == 766:
            for prime_mover, subset in plant.groupby("Reported Prime Mover"):
                net_mwh = subset["net_mwh"].sum()
                if net_mwh > 0:
                    rates[(plant_id, str(prime_mover))] = subset["fuel_mmbtu"].sum() / net_mwh
        else:
            net_mwh = plant["net_mwh"].sum()
            if net_mwh > 0:
                rates[(plant_id, "ALL")] = plant["fuel_mmbtu"].sum() / net_mwh

    return rates


def heco_2024_fuel_prices() -> dict[str, float]:
    workbook = load_workbook(HECO_WB3, read_only=True, data_only=True)
    sheet = workbook["EIA Fuel Price - Base Forecast"]
    row_2024 = next(row for row in sheet.iter_rows(values_only=True) if row[1] == 2024)
    return {fuel: float(row_2024[column - 1]) for fuel, column in FUEL_PRICE_COLUMN.items()}


def classify(plant_id: int, prime_mover: str) -> tuple[str, str, float, float]:
    """Return fuel, technology class, minimum fraction, and hourly ramp fraction."""
    if plant_id in {765}:
        return "LSFO", "steam", 0.40, 0.30
    if plant_id == 766 and prime_mover == "ST":
        return "LSFO", "steam", 0.40, 0.30
    if plant_id == 766:
        return "Diesel", "combustion_turbine", 0.0, 1.0
    if plant_id == 10334:
        return "Municipal solid waste", "municipal_waste", 0.80, 0.20
    if plant_id == 54646:
        return "LSFO", "combined_cycle", 0.40, 0.50
    if plant_id == 56329:
        return "ULSD_CIP", "combustion_turbine", 0.0, 1.0
    if plant_id == 60328:
        return "ULSD_SGS", "internal_combustion", 0.0, 1.0
    raise ValueError(f"No classification for plant {plant_id}")


def make_fleet() -> pd.DataFrame:
    generators = pd.read_excel(
        EIA860,
        sheet_name="Operable",
        header=1,
        dtype={"Generator ID": str},
    )
    generators = generators[
        generators["Plant Code"].isin(PLANT_CODES)
        & generators["State"].eq("HI")
        & generators["Status"].isin(["OP", "OS"])
        & ~generators["Prime Mover"].isin(["PV", "BA", "WT"])
    ].copy()
    generators["Plant Code"] = generators["Plant Code"].astype(int)
    generators["Summer Capacity (MW)"] = numeric(generators["Summer Capacity (MW)"])

    # Kalaeloa's two CTs and steam component form one combined-cycle resource; dispatching
    # the steam component independently would be physically incorrect.
    kalaeloa = generators[generators["Plant Code"].eq(54646)]
    generators = generators[~generators["Plant Code"].eq(54646)]

    rates = eia923_heat_rates()
    fuel_prices = heco_2024_fuel_prices()
    rows: list[dict[str, object]] = []

    for row in generators.to_dict("records"):
        plant_id = int(row["Plant Code"])
        prime_mover = str(row["Prime Mover"])
        fuel, technology, minimum_fraction, ramp_fraction = classify(plant_id, prime_mover)
        p_max = float(row["Summer Capacity (MW)"])
        heat_rate = rates[(plant_id, prime_mover if plant_id == 766 else "ALL")]
        if fuel == "Municipal solid waste":
            variable_cost = VOM_USD_PER_MWH[technology]
            co2_kg_per_mwh = 1489.548 * 0.45359237
            varcost_source = "EIA923-2024 heat rate; assumed zero fuel cost; ATB VOM proxy"
            emissions_source = "EPA-eGRID2023 HIOA output-rate proxy"
        else:
            variable_cost = heat_rate * fuel_prices[fuel] + VOM_USD_PER_MWH[technology]
            co2_kg_per_mwh = heat_rate * CO2_TONNES_PER_MMBTU[fuel] * 1000.0
            varcost_source = "EIA923-2024 heat rate; HECO-IGP-2022 2024 fuel forecast; ATB VOM proxy"
            emissions_source = "HECO-IGP-2022 fuel CO2 factor × EIA923-2024 heat rate"

        rows.append(
            {
                "name": f"{row['Plant Name']} {row['Generator ID']}",
                "fuel": fuel,
                "p_min_mw": p_max * minimum_fraction,
                "p_max_mw": p_max,
                "varcost_usd_per_mwh": variable_cost,
                "ramp_mw_per_hr": p_max * ramp_fraction,
                "source": "EIA-860/923 + HECO IGP; assumed operating constraints",
                "plant_code": plant_id,
                "generator_id": row["Generator ID"],
                "technology": technology,
                "heat_rate_mmbtu_per_mwh": heat_rate,
                "co2_kg_per_mwh": co2_kg_per_mwh,
                "capacity_source": "EIA860-2024 summer capacity",
                "varcost_source": varcost_source,
                "constraint_source": (
                    f"assumed: p_min={minimum_fraction:.0%} and ramp={ramp_fraction:.0%} of p_max per hour"
                ),
                "emissions_source": emissions_source,
            }
        )

    if not kalaeloa.empty:
        p_max = float(numeric(kalaeloa["Summer Capacity (MW)"]).sum())
        heat_rate = rates[(54646, "ALL")]
        fuel, technology, minimum_fraction, ramp_fraction = classify(54646, "CT")
        rows.append(
            {
                "name": "Kalaeloa combined cycle",
                "fuel": fuel,
                "p_min_mw": p_max * minimum_fraction,
                "p_max_mw": p_max,
                "varcost_usd_per_mwh": heat_rate * fuel_prices[fuel] + VOM_USD_PER_MWH[technology],
                "ramp_mw_per_hr": p_max * ramp_fraction,
                "source": "EIA-860/923 + HECO IGP; assumed operating constraints",
                "plant_code": 54646,
                "generator_id": "CT1+CT2+ST",
                "technology": technology,
                "heat_rate_mmbtu_per_mwh": heat_rate,
                "co2_kg_per_mwh": heat_rate * CO2_TONNES_PER_MMBTU[fuel] * 1000.0,
                "capacity_source": "EIA860-2024 sum of combined-cycle component summer capacities",
                "varcost_source": "EIA923-2024 plant heat rate; HECO-IGP-2022 2024 fuel forecast; ATB VOM proxy",
                "constraint_source": "assumed: p_min=40% and ramp=50% of p_max per hour",
                "emissions_source": "HECO-IGP-2022 fuel CO2 factor × EIA923-2024 heat rate",
            }
        )

    fleet = pd.DataFrame(rows).sort_values(["varcost_usd_per_mwh", "name"]).reset_index(drop=True)
    numeric_columns = [
        "p_min_mw",
        "p_max_mw",
        "varcost_usd_per_mwh",
        "ramp_mw_per_hr",
        "heat_rate_mmbtu_per_mwh",
        "co2_kg_per_mwh",
    ]
    fleet[numeric_columns] = fleet[numeric_columns].round(6)
    return fleet


def main() -> None:
    for path in [EIA860, EIA923, HECO_WB3]:
        if not path.exists():
            raise FileNotFoundError(f"Missing {path}; run scripts/download_data.py first")

    fleet = make_fleet()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    fleet.to_csv(OUTPUT, index=False)

    egrid = pd.DataFrame(
        [
            {
                "subregion": "HIOA",
                "name": "HICC Oahu",
                "data_year": 2023,
                "co2_lb_per_mwh": 1489.548,
                "co2_kg_per_mwh": 1489.548 * 0.45359237,
                "source": "EPA eGRID2023 Summary Tables, revision 2",
            }
        ]
    )
    egrid.round(6).to_csv(EGRID_OUTPUT, index=False)
    print(
        f"Wrote {len(fleet)} dispatchable resources with "
        f"{fleet.p_max_mw.sum():.1f} MW total summer capacity."
    )


if __name__ == "__main__":
    main()

