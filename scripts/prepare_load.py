#!/usr/bin/env python3
"""Extract the 2021 Oʻahu base-scenario hourly load layers from HECO Workbook 3."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data/raw/heco_20220331_final_oahu_inputs_workbook_3_revised.xlsx"
OUTPUT = ROOT / "data/processed/oahu_load_8760.csv"
COMPONENT_OUTPUT = ROOT / "data/processed/oahu_load_components_2021.csv"
YEAR = 2021

LAYER_SHEETS = {
    "underlying_load_mw": "Underlying_Load",
    "dgpv_mw": "DGPV Base Forecast",
    "residential_dbess_mw": "SchR_DBESS_Base",
    "commercial_dbess_mw": "SchG_DBESS_Base",
    "industrial_dbess_mw": "SchJ_DBESS_Base",
    "other_dbess_mw": "SchP_DBESS_Base",
    "managed_ev_mw": "Managed EV - Base Forecast",
    "energy_efficiency_mw": "EE - Base Forecast",
    "tou_shift_mw": "NON-DER-EV TOU - Base Forecast",
    "future_ebus_mw": "Future_eBus_Load",
}


def extract_layer(workbook, sheet_name: str) -> list[tuple[datetime, float]]:
    """Read a single HECO daily-by-hour layer and return hourly-ending values."""
    sheet = workbook[sheet_name]
    records: list[tuple[datetime, float]] = []
    started = False

    for row in sheet.iter_rows(values_only=True):
        date_value = row[1] if len(row) > 1 else None
        if not isinstance(date_value, datetime):
            continue
        if date_value.year < YEAR:
            continue
        if date_value.year > YEAR:
            break
        started = True
        for ending_hour, value in enumerate(row[2:26], start=1):
            timestamp = date_value + pd.Timedelta(hours=ending_hour)
            records.append((timestamp, float(value or 0.0)))

    if not started or len(records) != 8760:
        raise ValueError(f"{sheet_name}: expected 8,760 values for {YEAR}, found {len(records):,}")
    return records


def main() -> None:
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Missing {WORKBOOK}; run scripts/download_data.py first")

    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    frame: pd.DataFrame | None = None

    for column, sheet in LAYER_SHEETS.items():
        records = extract_layer(workbook, sheet)
        layer = pd.DataFrame(records, columns=["timestamp_hst", column])
        frame = layer if frame is None else frame.merge(layer, on="timestamp_hst", validate="one_to_one")

    assert frame is not None
    frame.insert(0, "hour", range(1, len(frame) + 1))
    frame["gross_load_mw"] = (
        frame["underlying_load_mw"] + frame["managed_ev_mw"] + frame["future_ebus_mw"]
    )
    adjustment_columns = [
        "dgpv_mw",
        "residential_dbess_mw",
        "commercial_dbess_mw",
        "industrial_dbess_mw",
        "other_dbess_mw",
        "energy_efficiency_mw",
        "tou_shift_mw",
    ]
    frame["net_load_mw"] = frame["gross_load_mw"] + frame[adjustment_columns].sum(axis=1)

    if frame[["gross_load_mw", "net_load_mw"]].isna().any().any():
        raise ValueError("Load extraction produced missing values")
    if (frame["net_load_mw"] <= 0).any():
        raise ValueError("Load extraction produced non-positive net load")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    frame[["hour", "gross_load_mw", "net_load_mw"]].round(6).to_csv(OUTPUT, index=False)
    frame.round(6).to_csv(COMPONENT_OUTPUT, index=False, date_format="%Y-%m-%dT%H:%M:%S-10:00")

    peak = frame.loc[frame["net_load_mw"].idxmax()]
    print(
        f"Wrote {len(frame):,} Oʻahu hours; net peak {peak.net_load_mw:.1f} MW "
        f"at hour {int(peak.hour):,}."
    )


if __name__ == "__main__":
    main()
